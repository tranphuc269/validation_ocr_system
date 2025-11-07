import json
from pathlib import Path
from datetime import datetime
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Request
from fastapi.responses import StreamingResponse
from pymongo.database import Database
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from ..config import settings
from ..database import get_db
from .. import crud, schemas
from ..utils.validation import extract_text_fields
from ..storage import storage_service


router = APIRouter()


@router.post("/", response_model=schemas.DocumentOut)
def create_document(payload: schemas.DocumentCreate, db: Database = Depends(get_db)):
    project = crud.get_project(db, str(payload.project_id)) if isinstance(payload.project_id, str) else crud.get_project(db, payload.project_id)  # type: ignore
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    return crud.create_document(db, payload)


@router.patch("/{doc_id}", response_model=schemas.DocumentOut)
def update_document(doc_id: str, payload: schemas.DocumentUpdate, db: Database = Depends(get_db)):
    doc = crud.update_document(db, doc_id, payload)
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    return doc


@router.get("/", response_model=list[schemas.DocumentOut])
def list_documents(project_id: str | None = None, db: Database = Depends(get_db)):
    return crud.list_documents(db, project_id)


@router.get("/{doc_id}", response_model=schemas.DocumentOut)
def get_document(doc_id: str, db: Database = Depends(get_db)):
    doc = crud.get_document_raw(db, doc_id)
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    return crud.serialize_document(doc)  # type: ignore


@router.delete("/{doc_id}")
def delete_document(doc_id: str, db: Database = Depends(get_db)):
    doc = crud.get_document_raw(db, doc_id)
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    
    # Delete sample JSON if exists
    if doc.get("sample_json_path"):
        try:
            storage_service.delete_file(doc["sample_json_path"])
        except Exception:
            pass
    
    # Get all uploads for this document and delete their files
    uploads = db["uploads"].find({"document_id": doc_id})
    for upload in uploads:
        # Delete uploaded file
        if upload.get("file_path"):
            try:
                storage_service.delete_file(upload["file_path"])
            except Exception:
                pass
        # Delete user input JSON if exists
        if upload.get("user_input_json_path"):
            try:
                storage_service.delete_file(upload["user_input_json_path"])
            except Exception:
                pass
    
    success = crud.delete_document(db, doc_id)
    if not success:
        raise HTTPException(status_code=404, detail="Document not found")
    return {"message": "Document deleted successfully"}


@router.post("/{doc_id}/sample-json", response_model=schemas.DocumentOut)
async def upload_sample_json(doc_id: str, sample: UploadFile = File(...), db: Database = Depends(get_db)):
    if not sample.filename.endswith(".json"):
        raise HTTPException(status_code=400, detail="Sample must be a JSON file")
    doc = crud.get_document_raw(db, doc_id)
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    content = await sample.read()
    # Validate JSON
    try:
        _ = json.loads(content.decode("utf-8"))
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON content")

    key = f"samples/doc_{doc_id}_sample.json"
    identifier = storage_service.save_bytes(key, content, content_type="application/json")
    updated = crud.set_document_sample_json_path(db, doc_id, identifier)
    if not updated:
        raise HTTPException(status_code=404, detail="Document not found")
    serialized = crud.serialize_document(updated)
    if not serialized:
        raise HTTPException(status_code=500, detail="Failed to serialize document")
    return serialized


@router.get("/{doc_id}/text-fields", response_model=list[str])
def get_text_fields(doc_id: str, db: Database = Depends(get_db)):
    doc = crud.get_document_raw(db, doc_id)
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    sample_path = doc.get("sample_json_path")
    if not sample_path:
        raise HTTPException(status_code=400, detail="Sample JSON not uploaded for document")
    try:
        parsed = json.loads(storage_service.read_text(sample_path))
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid sample JSON content")
    fields_map = extract_text_fields(parsed)
    return list(fields_map.keys())


@router.post("/{doc_id}/upload", response_model=schemas.UploadOut)
async def upload_file(doc_id: str, file: UploadFile = File(...), db: Database = Depends(get_db)):
    doc = crud.get_document_raw(db, doc_id)
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    ext = Path(file.filename).suffix.lower()
    if ext not in {".pdf", ".png", ".jpg", ".jpeg", ".tif", ".tiff"}:
        raise HTTPException(status_code=400, detail="Unsupported file type")
    content = await file.read()
    key = f"uploads/doc_{doc_id}_{file.filename}"
    identifier = storage_service.save_bytes(key, content, content_type=file.content_type or "application/octet-stream")
    return crud.create_upload(db, doc_id, identifier)


@router.post("/{upload_id}/user-input", response_model=schemas.UploadOut)
async def upload_user_input(upload_id: str, form_json: UploadFile = File(...), db: Database = Depends(get_db)):
    if not form_json.filename.endswith(".json"):
        raise HTTPException(status_code=400, detail="User input must be a JSON file")
    content = await form_json.read()
    # Validate JSON
    try:
        _ = json.loads(content.decode("utf-8"))
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON content")
    key = f"user_inputs/upload_{upload_id}_user.json"
    identifier = storage_service.save_bytes(key, content, content_type="application/json")
    updated = crud.set_upload_user_input(db, upload_id, identifier)
    if not updated:
        raise HTTPException(status_code=404, detail="Upload not found")
    serialized = crud.serialize_upload(updated)
    if not serialized:
        raise HTTPException(status_code=500, detail="Failed to serialize upload")
    return serialized


@router.get("/{doc_id}/uploads", response_model=list[schemas.UploadOut])
def list_uploads(doc_id: str, db: Database = Depends(get_db)):
    items = crud.list_uploads_by_document(db, doc_id)
    # serialize
    out: list[schemas.UploadOut] = []
    for it in items:
        out.append(
            schemas.UploadOut(
                id=str(it.get("_id")),
                document_id=it.get("document_id"),
                file_path=it.get("file_path"),
                user_input_json_path=it.get("user_input_json_path"),
                created_at=it.get("created_at"),
            )
        )
    return out


@router.get("/upload/{upload_id}", response_model=schemas.UploadOut)
def get_upload(upload_id: str, db: Database = Depends(get_db)):
    upload = crud.get_upload_raw(db, upload_id)
    if not upload:
        raise HTTPException(status_code=404, detail="Upload not found")
    return schemas.UploadOut(
        id=str(upload.get("_id")),
        document_id=upload.get("document_id"),
        file_path=upload.get("file_path"),
        user_input_json_path=upload.get("user_input_json_path"),
        created_at=upload.get("created_at"),
    )


@router.delete("/upload/{upload_id}")
def delete_upload(upload_id: str, db: Database = Depends(get_db)):
    upload = crud.get_upload_raw(db, upload_id)
    if not upload:
        raise HTTPException(status_code=404, detail="Upload not found")
    
    # Delete uploaded file
    if upload.get("file_path"):
        try:
            storage_service.delete_file(upload["file_path"])
        except Exception:
            pass
    
    # Delete user input JSON if exists
    if upload.get("user_input_json_path"):
        try:
            storage_service.delete_file(upload["user_input_json_path"])
        except Exception:
            pass
    
    success = crud.delete_upload(db, upload_id)
    if not success:
        raise HTTPException(status_code=404, detail="Upload not found")
    return {"message": "Upload deleted successfully"}


@router.get("/upload/{upload_id}/user-input", response_model=dict[str, str])
def get_upload_user_input(upload_id: str, db: Database = Depends(get_db)):
    upload = db["uploads"].find_one({"_id": crud._oid(upload_id)})
    if not upload:
        raise HTTPException(status_code=404, detail="Upload not found")
    path = upload.get("user_input_json_path")
    if not path:
        raise HTTPException(status_code=404, detail="User input not uploaded")
    try:
        data = json.loads(storage_service.read_text(path))
        if not isinstance(data, dict):
            raise ValueError("Invalid JSON structure")
        normalized: dict[str, str] = {}
        for key, value in data.items():
            normalized[str(key)] = "" if value is None else str(value)
        return normalized
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid user input JSON content")


@router.get("/upload/{upload_id}/file", name="get_upload_file")
def get_upload_file(upload_id: str, db: Database = Depends(get_db)):
    upload = crud.get_upload_raw(db, upload_id)
    if not upload:
        raise HTTPException(status_code=404, detail="Upload not found")
    file_path = upload.get("file_path")
    if not file_path:
        raise HTTPException(status_code=404, detail="File not found")
    identifier = str(file_path)
    name = identifier.split("/")[-1] if identifier else "download"
    return storage_service.file_response(identifier, download_name=name)


@router.get("/{doc_id}/export-excel")
def export_excel_report(doc_id: str, request: Request, db: Database = Depends(get_db)):
    """Export validation report to Excel for a document"""
    doc = crud.get_document_raw(db, doc_id)
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    
    # Get all uploads
    uploads = crud.list_uploads_by_document(db, doc_id)
    if not uploads:
        raise HTTPException(status_code=400, detail="No uploads found for this document")
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    
    # Sheet 1: Summary
    ws_summary = wb.create_sheet("Summary")
    ws_summary.append(["File Name", "Upload Date", "Overall Accuracy", "Total Fields", "Status", "Download Link"])
    
    # Style header
    for cell in ws_summary[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    summary_data = []
    all_details = []
    
    for upload in uploads:
        upload_id = str(upload.get("_id"))
        file_name = Path(upload.get("file_path", "")).name
        upload_date = upload.get("created_at")
        
        # Get validation results
        results = list(db["validation_results"].find({"upload_id": upload_id}).sort("created_at", -1))
        
        if results:
            overall_accuracy = sum(r.get("accuracy", 0) for r in results) / len(results) if results else 0.0
            status = "Validated"
        else:
            overall_accuracy = 0.0
            status = "Not Validated"
        
        download_url = str(request.url_for("get_upload_file", upload_id=upload_id))
        summary_data.append({
            "file_name": file_name,
            "upload_date": upload_date,
            "overall_accuracy": overall_accuracy,
            "total_fields": len(results),
            "status": status,
            "upload_id": upload_id,
            "download_url": download_url,
        })
        
        # Collect details
        for r in results:
            all_details.append({
                "file_name": file_name,
                "upload_date": upload_date,
                "field_name": r.get("field_name"),
                "user_value": r.get("user_value"),
                "ocr_value": r.get("ocr_value"),
                "accuracy": r.get("accuracy", 0.0),
            })
    
    # Write summary data
    for row_data in summary_data:
        ws_summary.append([
            row_data["file_name"],
            row_data["upload_date"].strftime("%Y-%m-%d %H:%M:%S") if row_data["upload_date"] else "N/A",
            f"{row_data['overall_accuracy'] * 100:.2f}%",
            row_data["total_fields"],
            row_data["status"],
            row_data["download_url"],
        ])
        row_idx = ws_summary.max_row
        link_cell = ws_summary.cell(row=row_idx, column=6)
        link_cell.hyperlink = row_data["download_url"]
        link_cell.value = "Download"
        link_cell.style = "Hyperlink"
    
    # Style summary rows
    for row in ws_summary.iter_rows(min_row=2, max_row=ws_summary.max_row):
        for cell in row:
            cell.border = border
            if cell.column == 3:  # Accuracy column
                cell.alignment = center_align
                accuracy_val = float(cell.value.replace('%', '')) / 100 if cell.value else 0
                if accuracy_val >= 0.9:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif accuracy_val >= 0.7:
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            if cell.column == 6:
                cell.alignment = center_align
    
    # Auto-adjust column widths for summary
    for column in ws_summary.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_summary.column_dimensions[column_letter].width = adjusted_width
    
    # Sheet 2: Detailed Results
    ws_details = wb.create_sheet("Detailed Results")
    ws_details.append(["File Name", "Upload Date", "Field Name", "User Value", "OCR Value", "Accuracy"])
    
    # Style header
    for cell in ws_details[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
    
    # Write details
    for detail in all_details:
        ws_details.append([
            detail["file_name"],
            detail["upload_date"].strftime("%Y-%m-%d %H:%M:%S") if detail["upload_date"] else "N/A",
            detail["field_name"],
            detail["user_value"],
            detail["ocr_value"],
            f"{detail['accuracy'] * 100:.2f}%",
        ])
    
    # Style details rows
    for row in ws_details.iter_rows(min_row=2, max_row=ws_details.max_row):
        for cell in row:
            cell.border = border
            if cell.column == 6:  # Accuracy column
                cell.alignment = center_align
                accuracy_val = float(cell.value.replace('%', '')) / 100 if cell.value else 0
                if accuracy_val >= 0.9:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif accuracy_val >= 0.7:
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Auto-adjust column widths for details
    for column in ws_details.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_details.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename
    doc_name = doc.get("name", "document").replace(" ", "_")
    filename = f"{doc_name}_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


